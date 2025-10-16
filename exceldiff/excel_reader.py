"""Excel file reader implementation."""

import os
from typing import List, Any, Optional
from openpyxl import load_workbook
from exceldiff.reader import FileReader


class ExcelReader(FileReader):
    """Reader for Excel (.xlsx) files."""

    def read(self, file_path: str, sheet_name: Optional[str] = None) -> List[List[Any]]:
        """
        Read a worksheet from an Excel file.

        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to read (None for first sheet)

        Returns:
            List of rows, where each row is a list of cell values
        """
        if not self.supports(file_path):
            raise ValueError(f"File {file_path} is not a valid .xlsx file")

        workbook = load_workbook(filename=file_path, data_only=True)

        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}"
                )
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active

        rows = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append(list(row))

        workbook.close()
        return rows

    def get_sheet_names(self, file_path: str) -> List[str]:
        """
        Get list of sheet names in the Excel file.

        Args:
            file_path: Path to the Excel file

        Returns:
            List of sheet names
        """
        if not self.supports(file_path):
            raise ValueError(f"File {file_path} is not a valid .xlsx file")

        workbook = load_workbook(filename=file_path, read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        return sheet_names

    def supports(self, file_path: str) -> bool:
        """
        Check if this reader supports the given file.

        Args:
            file_path: Path to the file

        Returns:
            True if the file has .xlsx extension
        """
        return os.path.splitext(file_path)[1].lower() == '.xlsx'
