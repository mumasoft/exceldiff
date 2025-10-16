"""Excel writer with color formatting for diffs."""

from typing import List
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
from exceldiff.differ import RowDiff, DiffType


class ExcelDiffWriter:
    """Writer for creating Excel files with diff highlighting."""

    # Color codes for different diff types
    COLOR_MODIFIED = "FF0000"  # Red for modified cells (new value)
    COLOR_ORIGINAL = "999900"  # Dark yellow for original values
    COLOR_REMOVED = "FFFF00"   # Yellow for removed rows
    COLOR_ADDED = "FFA500"     # Orange for added rows

    def write(self, diffs: List[RowDiff], output_path: str, diff_only: bool = False, include_header: bool = False) -> None:
        """
        Write diff results to an Excel file with color highlighting.

        Args:
            diffs: List of RowDiff objects
            output_path: Path to write the output file
            diff_only: If True, only write rows with differences (exclude identical rows)
            include_header: If True, include the first row as header (only applies when diff_only=True)

        Color scheme:
        - Identical rows: No coloring
        - Modified rows: Red cells for changed values
        - Removed rows: Yellow background for entire row
        - Added rows: Orange background for entire row
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Diff"

        # Filter diffs if needed
        diffs_to_write = [d for d in diffs if not diff_only or d.diff_type != DiffType.IDENTICAL]

        # Include header row if requested
        header_row = None
        if include_header and len(diffs) > 0:
            # First row is assumed to be the header
            header_row = diffs[0]

        # Write header if needed
        row_idx = 1
        if header_row:
            for col_idx, value in enumerate(header_row.row_data, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1

        # Write all rows with appropriate formatting
        for diff in diffs_to_write:
            # Write row data
            for col_idx, value in enumerate(diff.row_data, start=1):
                cell_idx = col_idx - 1

                # Apply formatting based on diff type
                if diff.diff_type == DiffType.IDENTICAL:
                    # No coloring for identical rows
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)

                elif diff.diff_type == DiffType.MODIFIED:
                    # For modified cells, show both old and new values
                    if cell_idx in diff.modified_cells and diff.original_row_data:
                        old_value = diff.original_row_data[cell_idx] if cell_idx < len(diff.original_row_data) else None
                        new_value = value

                        # Write cell with old and new values as plain text
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        old_str = str(old_value) if old_value is not None else ""
                        new_str = str(new_value) if new_value is not None else ""
                        cell.value = f"{old_str} → {new_str}"

                        # Apply red font color to indicate change
                        cell.font = Font(color=self.COLOR_MODIFIED)

                        # Add comment to show what changed
                        cell.comment = Comment(f"Changed from: {old_str}\nTo: {new_str}", "ExcelDiff")
                    else:
                        # Cell not modified, just write the value
                        cell = worksheet.cell(row=row_idx, column=col_idx, value=value)

                elif diff.diff_type == DiffType.REMOVED:
                    # Color entire row yellow
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.fill = PatternFill(
                        start_color=self.COLOR_REMOVED,
                        end_color=self.COLOR_REMOVED,
                        fill_type="solid"
                    )

                elif diff.diff_type == DiffType.ADDED:
                    # Color entire row orange
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.fill = PatternFill(
                        start_color=self.COLOR_ADDED,
                        end_color=self.COLOR_ADDED,
                        fill_type="solid"
                    )

            row_idx += 1

        # Auto-adjust column widths for better readability
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook
        workbook.save(output_path)
